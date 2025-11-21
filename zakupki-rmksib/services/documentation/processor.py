"""Сервис для обработки конкурсной документации"""
import os
import logging
from pathlib import Path
from typing import Optional
import aiofiles
import httpx
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse

logger = logging.getLogger(__name__)

# Поддерживаемые форматы
SUPPORTED_EXTENSIONS = {'.pdf', '.docx', '.doc', '.txt', '.rtf', '.xlsx', '.xls'}

# Magic bytes для определения типа файла
MAGIC_BYTES = {
	b'%PDF': '.pdf',
	b'PK\x03\x04': '.docx',  # ZIP-based formats (DOCX, XLSX)
	b'PK\x05\x06': '.docx',  # ZIP-based formats
	b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1': '.doc',  # Old MS Office formats
	b'\x50\x4b\x03\x04': '.xlsx',  # Excel (ZIP-based)
	b'{\\rtf': '.rtf',
	b'{\rtf': '.rtf',
}


async def save_documentation_file(file_content: bytes, filename: str, lot_number: str = None) -> str:
	"""
	Сохраняет файл документации на диск
	
	Args:
		file_content: Содержимое файла в байтах
		filename: Имя файла
		lot_number: Номер лота для организации структуры папок (опционально)
	
	Returns:
		Путь к сохраненному файлу
	"""
	# Создаем директорию для документации
	if lot_number:
		docs_dir = Path("data/documentation") / lot_number
	else:
		# Для файлов без привязки к лоту используем папку "manual"
		docs_dir = Path("data/documentation") / "manual"
	docs_dir.mkdir(parents=True, exist_ok=True)
	
	# Сохраняем файл
	file_path = docs_dir / filename
	async with aiofiles.open(file_path, 'wb') as f:
		await f.write(file_content)
	
	logger.info(f"Documentation file saved: {file_path}")
	return str(file_path)


async def extract_text_from_file(file_path: str) -> Optional[str]:
	"""
	Извлекает текст из файла документации
	
	Args:
		file_path: Путь к файлу
	
	Returns:
		Извлеченный текст или None в случае ошибки
	"""
	file_ext = Path(file_path).suffix.lower()
	
	# Если расширение не определено, пытаемся определить по содержимому
	if not file_ext or file_ext not in SUPPORTED_EXTENSIONS:
		logger.info(f"File extension not found or unsupported: {file_ext}, trying to detect by content")
		try:
			async with aiofiles.open(file_path, 'rb') as f:
				file_content = await f.read(1024)  # Читаем первые 1024 байта для определения типа
			
			detected_ext = detect_file_type_by_content(file_content)
			if detected_ext:
				logger.info(f"Detected file type by content: {detected_ext}")
				file_ext = detected_ext
				# Переименовываем файл, добавляя правильное расширение
				new_file_path = str(Path(file_path).with_suffix(detected_ext))
				if new_file_path != file_path:
					import shutil
					shutil.move(file_path, new_file_path)
					file_path = new_file_path
					logger.info(f"Renamed file to: {new_file_path}")
			else:
				logger.warning(f"Could not detect file type for: {file_path}")
				return None
		except Exception as e:
			logger.error(f"Error detecting file type for {file_path}: {e}")
			return None
	
	try:
		if file_ext == '.pdf':
			return await _extract_text_from_pdf(file_path)
		elif file_ext in {'.docx', '.doc'}:
			return await _extract_text_from_docx(file_path)
		elif file_ext == '.txt':
			return await _extract_text_from_txt(file_path)
		elif file_ext == '.rtf':
			return await _extract_text_from_rtf(file_path)
		elif file_ext in {'.xlsx', '.xls'}:
			return await _extract_text_from_excel(file_path)
		else:
			logger.warning(f"Unsupported file format: {file_ext}")
			return None
	except Exception as e:
		logger.error(f"Error extracting text from {file_path}: {e}")
		return None


async def _extract_text_from_pdf(file_path: str) -> str:
	"""Извлекает текст из PDF файла"""
	try:
		import PyPDF2
		import io
		text = ""
		async with aiofiles.open(file_path, 'rb') as f:
			file_content = await f.read()
			pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
			for page in pdf_reader.pages:
				text += page.extract_text() + "\n"
		return text.strip()
	except ImportError:
		logger.error("PyPDF2 not installed. Install it with: pip install PyPDF2")
		# Fallback: возвращаем сообщение об ошибке
		return f"[Ошибка: PyPDF2 не установлен. Установите: pip install PyPDF2]"
	except Exception as e:
		logger.error(f"Error reading PDF {file_path}: {e}")
		return f"[Ошибка при чтении PDF: {str(e)}]"


async def _extract_text_from_docx(file_path: str) -> str:
	"""Извлекает текст из DOCX файла"""
	try:
		from docx import Document
		doc = Document(file_path)
		text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
		return text.strip()
	except ImportError:
		logger.error("python-docx not installed. Install it with: pip install python-docx")
		return f"[Ошибка: python-docx не установлен. Установите: pip install python-docx]"
	except Exception as e:
		logger.error(f"Error reading DOCX {file_path}: {e}")
		return f"[Ошибка при чтении DOCX: {str(e)}]"


async def _extract_text_from_txt(file_path: str) -> str:
	"""Извлекает текст из TXT файла"""
	try:
		async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
			text = await f.read()
		return text.strip()
	except UnicodeDecodeError:
		# Пробуем другие кодировки
		try:
			async with aiofiles.open(file_path, 'r', encoding='cp1251') as f:
				text = await f.read()
			return text.strip()
		except Exception as e:
			logger.error(f"Error reading TXT {file_path}: {e}")
			return f"[Ошибка при чтении TXT: {str(e)}]"
	except Exception as e:
		logger.error(f"Error reading TXT {file_path}: {e}")
		return f"[Ошибка при чтении TXT: {str(e)}]"


async def _extract_text_from_rtf(file_path: str) -> str:
	"""Извлекает текст из RTF файла (базовая реализация)"""
	try:
		import re
		async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
			content = await f.read()
		# Простое удаление RTF-тегов (базовая реализация)
		text = re.sub(r'\{[^}]*\}', '', content)
		text = re.sub(r'\\[a-z]+\d*\s?', '', text)
		return text.strip()
	except Exception as e:
		logger.error(f"Error reading RTF {file_path}: {e}")
		return f"[Ошибка при чтении RTF: {str(e)}]"


async def _extract_text_from_excel(file_path: str) -> str:
	"""Извлекает текст из Excel файла"""
	try:
		import openpyxl
		text = ""
		workbook = openpyxl.load_workbook(file_path, data_only=True)
		
		# Обрабатываем все листы
		for sheet_name in workbook.sheetnames:
			sheet = workbook[sheet_name]
			text += f"\n\n=== Лист: {sheet_name} ===\n\n"
			
			# Читаем все ячейки с данными
			for row in sheet.iter_rows(values_only=True):
				row_text = []
				for cell_value in row:
					if cell_value is not None:
						row_text.append(str(cell_value))
				if row_text:
					text += " | ".join(row_text) + "\n"
		
		return text.strip()
	except ImportError:
		logger.error("openpyxl not installed. Install it with: pip install openpyxl")
		return f"[Ошибка: openpyxl не установлен. Установите: pip install openpyxl]"
	except Exception as e:
		logger.error(f"Error reading Excel {file_path}: {e}")
		return f"[Ошибка при чтении Excel: {str(e)}]"


def is_supported_format(filename: str) -> bool:
	"""Проверяет, поддерживается ли формат файла"""
	file_ext = Path(filename).suffix.lower()
	return file_ext in SUPPORTED_EXTENSIONS


def detect_file_type_by_content(file_content: bytes) -> Optional[str]:
	"""
	Определяет тип файла по содержимому (magic bytes)
	
	Args:
		file_content: Содержимое файла в байтах
		
	Returns:
		Расширение файла (например, '.pdf') или None
	"""
	if not file_content:
		return None
	
	# Проверяем magic bytes
	for magic, ext in MAGIC_BYTES.items():
		if file_content.startswith(magic):
			return ext
	
	# Для ZIP-based форматов (DOCX, XLSX) нужно проверить внутреннюю структуру
	if file_content.startswith(b'PK\x03\x04') or file_content.startswith(b'PK\x05\x06'):
		# Проверяем наличие типичных файлов в архиве (читаем больше для надежности)
		content_str = file_content[:2048]  # Первые 2048 байт для более надежного определения
		# Excel файлы содержат xl/ или worksheets/
		if b'xl/' in content_str or b'worksheets/' in content_str or b'workbook.xml' in content_str:
			return '.xlsx'
		# Word файлы содержат word/ или document.xml
		elif b'word/' in content_str or b'document.xml' in content_str or b'[Content_Types].xml' in content_str:
			return '.docx'
		# Если не удалось определить точно, но это ZIP - скорее всего DOCX
		else:
			return '.docx'
	
	# Если не удалось определить, возвращаем None
	return None


async def download_documentation_from_url(url: str, lot_number: str) -> Optional[str]:
	"""
	Скачивает документацию с URL страницы лота
	
	Args:
		url: URL страницы лота на площадке закупок
		lot_number: Номер лота для организации структуры папок
	
	Returns:
		Путь к сохраненному файлу документации или None в случае ошибки
	"""
	try:
		async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
			headers = {
				"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
			}
			
			# Загружаем страницу лота
			response = await client.get(url, headers=headers)
			response.raise_for_status()
			
			html = await response.aread()
			soup = BeautifulSoup(html, 'html.parser')
			
			# Ищем ссылки на документацию (обычно это ссылки на PDF, DOCX, DOC, XLSX файлы)
			doc_links = []
			
			# Ищем все ссылки с поддерживаемыми расширениями
			for link in soup.find_all('a', href=True):
				href = link.get('href', '')
				link_text = link.get_text(strip=True).lower()
				
				# Проверяем расширение файла
				parsed_url = urlparse(href)
				file_ext = Path(parsed_url.path).suffix.lower()
				
				# Если это прямая ссылка на файл
				if file_ext in SUPPORTED_EXTENSIONS:
					full_url = urljoin(url, href)
					doc_links.append((full_url, file_ext, link_text))
				
				# Также ищем по тексту ссылки (например, "скачать", "документация", "техническое задание")
				elif any(keyword in link_text for keyword in ['скачать', 'документация', 'техническое', 'задание', 'проект', 'документ']):
					full_url = urljoin(url, href)
					# Проверяем, может быть это страница с файлом
					if file_ext in SUPPORTED_EXTENSIONS or not file_ext:
						doc_links.append((full_url, file_ext, link_text))
			
			if not doc_links:
				logger.warning(f"No documentation links found on page: {url}")
				return None
			
			# Берем первую найденную ссылку (можно улучшить логику выбора)
			doc_url, file_ext, link_text = doc_links[0]
			
			# Если расширение не определено, пробуем определить по Content-Type
			if not file_ext:
				head_response = await client.head(doc_url, headers=headers)
				content_type = head_response.headers.get('Content-Type', '')
				if 'pdf' in content_type.lower():
					file_ext = '.pdf'
				elif 'word' in content_type.lower() or 'document' in content_type.lower():
					file_ext = '.docx'
				elif 'excel' in content_type.lower() or 'spreadsheet' in content_type.lower():
					file_ext = '.xlsx'
				else:
					file_ext = '.pdf'  # По умолчанию PDF
			
			# Скачиваем файл
			logger.info(f"Downloading documentation from: {doc_url}")
			file_response = await client.get(doc_url, headers=headers)
			file_response.raise_for_status()
			
			file_content = await file_response.aread()
			
			# Проверяем, не является ли ответ HTML страницей
			content_type = file_response.headers.get('Content-Type', '')
			if 'text/html' in content_type.lower() or file_content.startswith(b'<html') or file_content.startswith(b'<!DOCTYPE'):
				logger.warning(f"Downloaded content is HTML, not a file. URL: {doc_url}")
				# Пытаемся найти ссылки на файлы в HTML
				try:
					html_soup = BeautifulSoup(file_content, 'html.parser')
					# Ищем ссылки на файлы документации
					for link in html_soup.find_all('a', href=True):
						href = link.get('href', '')
						parsed_href = urlparse(href)
						href_ext = Path(parsed_href.path).suffix.lower()
						if href_ext in SUPPORTED_EXTENSIONS:
							# Нашли ссылку на файл, скачиваем его
							file_url = urljoin(doc_url, href)
							logger.info(f"Found file link in HTML: {file_url}")
							file_response = await client.get(file_url, headers=headers)
							file_response.raise_for_status()
							file_content = await file_response.aread()
							file_ext = href_ext
							doc_url = file_url  # Обновляем URL для логирования
							break
					else:
						# Не нашли ссылок на файлы
						logger.error(f"No file links found in HTML page: {doc_url}")
						return None
				except Exception as e:
					logger.error(f"Error parsing HTML page: {e}")
					return None
			
			# Определяем тип файла по содержимому, если не удалось определить ранее
			if not file_ext:
				detected_ext = detect_file_type_by_content(file_content)
				if detected_ext:
					file_ext = detected_ext
					logger.info(f"Detected file type by content: {file_ext}")
				else:
					# Если не удалось определить, пробуем по Content-Type из ответа
					content_type = file_response.headers.get('Content-Type', '')
					if 'pdf' in content_type.lower():
						file_ext = '.pdf'
					elif 'word' in content_type.lower() or 'document' in content_type.lower():
						file_ext = '.docx'
					elif 'excel' in content_type.lower() or 'spreadsheet' in content_type.lower():
						file_ext = '.xlsx'
					else:
						file_ext = '.pdf'  # По умолчанию PDF
					logger.info(f"Using Content-Type to determine file type: {file_ext}")
			
			# Генерируем имя файла
			parsed_doc_url = urlparse(doc_url)
			filename = Path(parsed_doc_url.path).name
			if not filename or filename == '/' or not Path(filename).suffix:
				# Если имя файла пустое или без расширения, создаем с правильным расширением
				filename = f"documentation{file_ext}"
			elif Path(filename).suffix.lower() not in SUPPORTED_EXTENSIONS:
				# Если расширение не поддерживается, заменяем на определенное
				filename = Path(filename).stem + file_ext
			
			# Сохраняем файл
			file_path = await save_documentation_file(file_content, filename, lot_number)
			
			logger.info(f"Documentation downloaded and saved: {file_path}")
			
			# Пытаемся автоматически извлечь текст из документации
			try:
				documentation_text = await extract_text_from_file(file_path)
				if documentation_text and not documentation_text.startswith("[Ошибка"):
					# Сохраняем извлеченный текст в БД
					from database.repositories.lot import LotRepository
					from database.connection import async_session_maker
					
					async with async_session_maker() as session:
						lot_repo = LotRepository(session)
						lot = await lot_repo.get_by_lot_number(lot_number)
						if lot:
							lot.documentation_text = documentation_text
							await lot_repo.update(lot)
							logger.info(f"Documentation text extracted and saved for lot {lot_number}: {len(documentation_text)} characters")
			except Exception as e:
				logger.warning(f"Could not extract text from documentation automatically: {e}")
				# Не критично, пользователь может извлечь текст позже
			
			return file_path
			
	except httpx.HTTPStatusError as e:
		logger.error(f"HTTP error downloading documentation from {url}: {e}")
		return None
	except Exception as e:
		logger.error(f"Error downloading documentation from {url}: {e}", exc_info=True)
		return None



