"""Сервис для парсинга Excel файлов и извлечения списка товаров"""
import logging
from typing import List, Dict, Optional
from pathlib import Path
from loguru import logger

logger = logging.getLogger(__name__)


async def extract_products_from_excel(file_path: str) -> List[Dict[str, Optional[str]]]:
    """
    Извлекает список товаров из Excel файла
    
    Args:
        file_path: Путь к файлу Excel
        
    Returns:
        Список словарей с информацией о товарах:
        [
            {
                "name": "Название товара",
                "code": "Код номенклатуры",
                "row_number": 15,
                "quantity": "Количество" (если найдено),
                "unit": "Единица измерения" (если найдено, например: шт., кг., м.)
            },
            ...
        ]
    """
    products = []
    
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Ищем заголовки таблицы (обычно в первых 20 строках)
        header_row = None
        name_col = None
        code_col = None
        quantity_col = None
        unit_col = None
        
        # Маркеры для поиска столбцов
        name_markers = ['номенклатура', 'наименование', 'товар', 'название', 'описание']
        code_markers = ['код номенклатуры', 'код', 'артикул', 'номер']
        quantity_markers = ['количество', 'кол-во', 'qty', 'шт', 'штук']
        unit_markers = ['единица', 'ед.', 'ед', 'единица измерения', 'единицы', 'измерения', 'размерность']
        
        # Ищем строку с заголовками
        for row_idx in range(1, min(21, sheet.max_row + 1)):
            row = sheet[row_idx]
            for col_idx, cell in enumerate(row, 1):
                cell_value = str(cell.value or '').lower().strip()
                
                # Ищем столбец с названием товара
                if not name_col:
                    for marker in name_markers:
                        if marker in cell_value:
                            name_col = col_idx
                            header_row = row_idx
                            break
                
                # Ищем столбец с кодом
                if not code_col:
                    for marker in code_markers:
                        if marker in cell_value:
                            code_col = col_idx
                            if not header_row:
                                header_row = row_idx
                            break
                
                # Ищем столбец с количеством
                if not quantity_col:
                    for marker in quantity_markers:
                        if marker in cell_value:
                            quantity_col = col_idx
                            if not header_row:
                                header_row = row_idx
                            break
                
                # Ищем столбец с единицами измерения
                if not unit_col:
                    for marker in unit_markers:
                        if marker in cell_value:
                            unit_col = col_idx
                            if not header_row:
                                header_row = row_idx
                            break
        
        if not header_row:
            logger.warning("Could not find header row in Excel file")
            # Пробуем найти товары без заголовков - ищем строки с данными
            # Обычно товары начинаются после строки 10
            header_row = 10
        
        # Если не нашли столбец с названием, пробуем найти по структуре
        if not name_col:
            # Ищем столбец с наибольшим количеством непустых значений (обычно это название товара)
            max_non_empty = 0
            for col_idx in range(1, min(sheet.max_column + 1, 20)):
                non_empty_count = 0
                for row_idx in range(header_row + 1, min(header_row + 20, sheet.max_row + 1)):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and str(cell_value).strip():
                        non_empty_count += 1
                if non_empty_count > max_non_empty:
                    max_non_empty = non_empty_count
                    name_col = col_idx
        
        if not name_col:
            logger.error("Could not determine name column in Excel file")
            workbook.close()
            return products
        
        # Извлекаем товары из строк после заголовка
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            name_cell = sheet.cell(row=row_idx, column=name_col)
            name_value = str(name_cell.value or '').strip()
            
            # Пропускаем пустые строки и служебные строки
            if not name_value or len(name_value) < 3:
                continue
            
            # Пропускаем строки с метаданными (например, "Итого", "Всего", "Ответственный")
            skip_keywords = ['итого', 'всего', 'ответственный', 'согласовано', 'должность', 
                           'контролирующий', 'склад', 'подразделение', 'сценарий', 'цфо']
            if any(keyword in name_value.lower() for keyword in skip_keywords):
                continue
            
            # Извлекаем код товара (если найден столбец)
            code_value = None
            if code_col:
                code_cell = sheet.cell(row=row_idx, column=code_col)
                code_value = str(code_cell.value or '').strip() if code_cell.value else None
            
            # Извлекаем количество (если найден столбец)
            quantity_value = None
            if quantity_col:
                quantity_cell = sheet.cell(row=row_idx, column=quantity_col)
                quantity_value = str(quantity_cell.value or '').strip() if quantity_cell.value else None
            
            # Извлекаем единицы измерения (если найден столбец)
            unit_value = None
            if unit_col:
                unit_cell = sheet.cell(row=row_idx, column=unit_col)
                unit_value = str(unit_cell.value or '').strip() if unit_cell.value else None
            
            # Если единицы измерения не найдены в отдельном столбце, пытаемся извлечь из количества
            if not unit_value and quantity_value:
                # Ищем единицы измерения в строке количества (например: "100 шт", "50 кг", "10 м")
                import re
                unit_pattern = r'\s*(шт|штук|кг|килограмм|г|грамм|т|тонн|м|метр|см|сантиметр|мм|миллиметр|л|литр|мл|миллилитр|м²|м2|м³|м3|упак|упаковок|компл|комплект|пар|пар|пог\.?\s*м|пог\.?\s*метр|п\.\s*м|п\.\s*метр)\.?$'
                unit_match = re.search(unit_pattern, quantity_value, re.IGNORECASE)
                if unit_match:
                    unit_value = unit_match.group(1).strip()
                    # Удаляем единицы измерения из количества
                    quantity_value = re.sub(unit_pattern, '', quantity_value, flags=re.IGNORECASE).strip()
            
            # Добавляем товар в список
            product = {
                "name": name_value,
                "code": code_value,
                "row_number": row_idx,
                "quantity": quantity_value,
                "unit": unit_value
            }
            products.append(product)
        
        workbook.close()
        logger.info(f"Extracted {len(products)} products from Excel file: {file_path}")
        
    except ImportError:
        logger.error("openpyxl not installed. Install it with: pip install openpyxl")
    except Exception as e:
        logger.error(f"Error extracting products from Excel {file_path}: {e}", exc_info=True)
    
    return products

